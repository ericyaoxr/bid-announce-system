import csv
import io
import json
import urllib.parse
from datetime import UTC, datetime

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from src.api.deps import CurrentUser, DbSession
from src.db.models import Announcement

router = APIRouter(prefix="/api/export", tags=["数据导出"])


@router.get("/formats")
async def get_export_formats(_user: CurrentUser) -> dict:
    return {"formats": ["csv", "excel"], "labels": {"csv": "CSV 文件", "excel": "Excel 文件"}}


async def _query_export_data(
    db: DbSession,
    keyword: str | None = None,
    category: str | None = None,
    tender_mode: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[dict]:
    query = select(Announcement).order_by(Announcement.publish_date.desc())

    if keyword:
        kw = f"%{keyword}%"
        query = query.where(
            (Announcement.title.ilike(kw))
            | (Announcement.project_no.ilike(kw))
            | (Announcement.tenderer_name.ilike(kw))
        )
    if category:
        query = query.where(Announcement.category == category)
    if tender_mode:
        query = query.where(Announcement.tender_mode_desc == tender_mode)
    if start_date:
        query = query.where(Announcement.publish_date >= start_date)
    if end_date:
        query = query.where(Announcement.publish_date <= end_date + "T23:59:59")

    result = await db.execute(query)
    rows = result.scalars().all()

    data: list[dict] = []
    for r in rows:
        wbs_all = []
        if r.winning_bidders and r.winning_bidders not in ("", "[]"):
            try:
                wbs_all = json.loads(r.winning_bidders)
            except (json.JSONDecodeError, TypeError):
                pass

        wbs = [w for w in wbs_all if w.get("is_winning") == 1]

        if not wbs:
            data.append(
                {
                    "id": r.id,
                    "project_no": r.project_no or "",
                    "title": r.title or "",
                    "category": r.category or "",
                    "tender_mode_desc": r.tender_mode_desc or "",
                    "tenderer_name": r.tenderer_name or "",
                    "tenderer_contact": r.tenderer_contact or "",
                    "tenderer_phone": r.tenderer_phone or "",
                    "publish_date": r.publish_date or "",
                    "project_address": r.project_address or "",
                    "fund_source": r.fund_source or "",
                    "purchase_control_price": r.purchase_control_price or "",
                    "bid_price": r.bid_price or "",
                    "winner_supplier": "",
                    "winner_amount": "",
                    "winner_credit_code": "",
                    "source_url": r.source_url or "",
                }
            )
        else:
            for w in wbs:
                data.append(
                    {
                        "id": r.id,
                        "project_no": r.project_no or "",
                        "title": r.title or "",
                        "category": r.category or "",
                        "tender_mode_desc": r.tender_mode_desc or "",
                        "tenderer_name": r.tenderer_name or "",
                        "tenderer_contact": r.tenderer_contact or "",
                        "tenderer_phone": r.tenderer_phone or "",
                        "publish_date": r.publish_date or "",
                        "project_address": r.project_address or "",
                        "fund_source": r.fund_source or "",
                        "purchase_control_price": r.purchase_control_price or "",
                        "bid_price": r.bid_price or "",
                        "winner_supplier": w.get("supplier_name", ""),
                        "winner_amount": w.get("bid_amount", ""),
                        "winner_credit_code": w.get("social_credit_code", ""),
                        "source_url": r.source_url or "",
                    }
                )
    return data


@router.get("/csv")
async def export_csv(
    db: DbSession,
    _user: CurrentUser,
    keyword: str | None = Query(None),
    category: str | None = Query(None),
    tender_mode: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
) -> StreamingResponse:
    data = await _query_export_data(db, keyword, category, tender_mode, start_date, end_date)

    output = io.StringIO()
    output.write("\ufeff")

    headers = [
        "公告ID",
        "项目编号",
        "公告标题",
        "分类",
        "招标方式",
        "招标人",
        "联系人",
        "联系电话",
        "发布时间",
        "项目地址",
        "资金来源",
        "控制价",
        "中标金额",
        "中标人",
        "中标人信用代码",
        "详情链接",
    ]
    writer = csv.writer(output)
    writer.writerow(headers)
    for d in data:
        writer.writerow(
            [
                d["id"],
                d["project_no"],
                d["title"],
                d["category"],
                d["tender_mode_desc"],
                d["tenderer_name"],
                d["tenderer_contact"],
                d["tenderer_phone"],
                d["publish_date"],
                d["project_address"],
                d["fund_source"],
                d["purchase_control_price"],
                d["winner_amount"],
                d["winner_supplier"],
                d["winner_credit_code"],
                d["source_url"],
            ]
        )

    filename = f"中标结果_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.csv"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/excel")
async def export_excel(
    db: DbSession,
    _user: CurrentUser,
    keyword: str | None = Query(None),
    category: str | None = Query(None),
    tender_mode: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = await _query_export_data(db, keyword, category, tender_mode, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "中标结果"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    amount_font = Font(name="微软雅黑", size=11, color="CC0000", bold=True)
    winner_font = Font(name="微软雅黑", size=11, color="1677FF")
    thin_border = Border(
        left=Side(style="thin", color="E8E8E8"),
        right=Side(style="thin", color="E8E8E8"),
        top=Side(style="thin", color="E8E8E8"),
        bottom=Side(style="thin", color="E8E8E8"),
    )

    headers = [
        ("公告ID", 10),
        ("项目编号", 18),
        ("公告标题", 45),
        ("分类", 8),
        ("招标方式", 12),
        ("招标人", 20),
        ("联系人", 10),
        ("联系电话", 14),
        ("发布时间", 18),
        ("项目地址", 25),
        ("资金来源", 15),
        ("控制价", 14),
        ("中标金额", 14),
        ("中标人", 22),
        ("中标人信用代码", 22),
        ("详情链接", 30),
    ]

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, d in enumerate(data, 2):
        values = [
            d["id"],
            d["project_no"],
            d["title"],
            d["category"],
            d["tender_mode_desc"],
            d["tenderer_name"],
            d["tenderer_contact"],
            d["tenderer_phone"],
            d["publish_date"],
            d["project_address"],
            d["fund_source"],
            d["purchase_control_price"],
            d["winner_amount"],
            d["winner_supplier"],
            d["winner_credit_code"],
            d["source_url"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            cell.alignment = cell_align
            cell.border = thin_border
            if col_idx in (12, 13) and val:
                cell.font = amount_font
                cell.number_format = "#,##0.00"
            if col_idx == 14 and val:
                cell.font = winner_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"中标结果_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
