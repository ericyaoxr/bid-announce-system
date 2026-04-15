"""FastAPI后端API v2 - 支持深度数据（中标人/金额）+ 真正的采集数据执行 + 导出"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import sqlite3
import threading
import time
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from src.utils.logger import get_logger

logger = get_logger(__name__)

app = FastAPI(
    title="中标结果公示系统 API",
    description="提供公告查询、统计、采集数据管理接口",
    version="2.1.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = str(Path(__file__).parent.parent.parent / "data" / "announcements_deep.db")


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ========== 采集数据任务管理 ==========

class CrawlerTask:
    """采集数据任务状态管理"""
    def __init__(self):
        self.is_running = False
        self.mode = ""
        self.task_id = ""
        self.start_time: float | None = None
        self.logs: list[dict] = []
        self.progress: dict = {}
        self.result: dict = {}
        self._stop_flag = False
        self._thread: threading.Thread | None = None

    def add_log(self, msg: str, level: str = "info"):
        self.logs.insert(0, {
            "time": datetime.now().strftime("%H:%M:%S"),
            "level": level,
            "msg": msg,
        })
        # 只保留最近200条
        if len(self.logs) > 200:
            self.logs = self.logs[:200]

    def to_dict(self) -> dict:
        elapsed = ""
        if self.start_time:
            if self.is_running:
                elapsed = f"{time.time() - self.start_time:.1f}s"
            else:
                elapsed = f"{self.result.get('elapsed', 0):.1f}s"

        return {
            "is_running": self.is_running,
            "mode": self.mode,
            "task_id": self.task_id,
            "elapsed": elapsed,
            "progress": self.progress,
            "result": self.result,
            "recent_logs": self.logs[:50],
            "log_count": len(self.logs),
        }


# 全局采集数据任务实例
crawler_task = CrawlerTask()


def _run_crawler_sync(mode: str, max_pages: int, days: int | None = None):
    """在线程中同步运行采集数据"""
    import asyncio

    crawler_task.is_running = True
    crawler_task.mode = mode
    crawler_task.start_time = time.time()
    crawler_task.progress = {"phase": "starting", "detail": "初始化..."}
    crawler_task.result = {}
    crawler_task._stop_flag = False

    mode_label = {"incremental": "增量", "full": "全量", "by_date": "按时间", "detail_only": "仅详情"}.get(mode, mode)
    extra = f", 最近{days}天" if mode == "by_date" and days else ""
    crawler_task.add_log(f"启动{mode_label}采集, max_pages={max_pages}{extra}")

    try:
        # 在新的事件循环中运行
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(_async_crawler_main(mode, max_pages, days))
        loop.close()
    except Exception as e:
        crawler_task.add_log(f"采集异常: {str(e)}", "error")
        crawler_task.result = {"status": "error", "error": str(e)}
    finally:
        elapsed = time.time() - crawler_task.start_time
        crawler_task.is_running = False
        crawler_task.result["elapsed"] = elapsed
        crawler_task.progress["phase"] = "done"
        crawler_task.add_log(f"采集完成, 耗时 {elapsed:.1f}s")


async def _async_crawler_main(mode: str, max_pages: int, days: int | None = None):
    """异步采集数据主逻辑 - 所有模式只采集结果公示(type=4)+有中标"""
    from src.crawlers.deep_crawler import DeepCrawler

    crawler = DeepCrawler(db_path=DB_PATH, rate_limit_rpm=120)

    try:
        if mode == "incremental":
            # 增量采集: 抓结果公示(type=4) + 增量模式(遇到旧数据停止) + 抓详情 + 删除无中标
            crawler_task.add_log("增量采集: 抓取结果公示新数据...")
            crawler_task.progress = {"phase": "list", "detail": "抓取结果公示新数据..."}

            list_count = await crawler.crawl_list(4, max_pages=max_pages, incremental=True)
            crawler_task.add_log(f"列表抓取完成: 新增 {list_count} 条")
            crawler_task.progress = {"phase": "detail", "detail": f"列表完成(新增{list_count}条), 开始抓详情...", "list_count": list_count}

            detail_count = await crawler.crawl_details(announcement_type=4)
            crawler_task.add_log(f"详情抓取完成: {detail_count} 条")

            # 删除无中标人的记录
            removed = crawler.remove_no_winner()
            crawler_task.add_log(f"清理无中标记录: 删除 {removed} 条")

            # 统计结果
            conn = sqlite3.connect(DB_PATH)
            total = conn.execute("SELECT COUNT(*) FROM announcements").fetchone()[0]
            with_winner = conn.execute("SELECT COUNT(*) FROM announcements WHERE winning_bidders IS NOT NULL AND winning_bidders != '[]'").fetchone()[0]
            conn.close()

            crawler_task.result = {
                "status": "success",
                "list_count": list_count,
                "detail_count": detail_count,
                "removed_no_winner": removed,
                "total_records": total,
                "with_winner": with_winner,
            }
            crawler_task.add_log(f"增量采集完成! 新增={list_count}, 总记录={total}, 有中标={with_winner}")

        elif mode == "full":
            # 全量采集: 抓结果公示(type=4)全部页 + 详情 + 删除无中标
            crawler_task.add_log("全量采集: 抓取结果公示全部数据...")
            crawler_task.progress = {"phase": "list", "detail": "抓取结果公示列表..."}

            list_count = await crawler.crawl_list(4, max_pages=max_pages)
            crawler_task.add_log(f"结果公示: 新增 {list_count} 条")

            if not crawler_task._stop_flag:
                crawler_task.add_log(f"列表抓取完成，共 {list_count} 条，开始抓详情...")
                crawler_task.progress = {"phase": "detail", "detail": f"列表完成({list_count}条), 开始抓详情...", "list_count": list_count}

                detail_count = await crawler.crawl_details(announcement_type=4)
                crawler_task.add_log(f"详情抓取完成: {detail_count} 条")

                # 删除无中标人的记录
                removed = crawler.remove_no_winner()
                crawler_task.add_log(f"清理无中标记录: 删除 {removed} 条")

                conn = sqlite3.connect(DB_PATH)
                total = conn.execute("SELECT COUNT(*) FROM announcements").fetchone()[0]
                with_winner = conn.execute("SELECT COUNT(*) FROM announcements WHERE winning_bidders IS NOT NULL AND winning_bidders != '[]'").fetchone()[0]
                conn.close()

                crawler_task.result = {
                    "status": "success",
                    "list_count": list_count,
                    "detail_count": detail_count,
                    "removed_no_winner": removed,
                    "total_records": total,
                    "with_winner": with_winner,
                }
                crawler_task.add_log(f"全量采集完成! 总记录={total}, 有中标={with_winner}")
            else:
                crawler_task.result = {"status": "stopped", "list_count": list_count}

        elif mode == "by_date":
            # 按时间范围采集: 抓结果公示(type=4)最近N天 + 详情 + 删除无中标
            actual_days = days or 30
            crawler_task.add_log(f"按时间采集: 结果公示最近 {actual_days} 天...")
            crawler_task.progress = {"phase": "list", "detail": f"按时间采集(最近{actual_days}天)..."}

            # 按时间采集：估算页数（假设每天约3-5条，每页20条）
            est_pages = max(1, (actual_days * 5 + 19) // 20)
            list_count = await crawler.crawl_list(4, max_pages=est_pages, incremental=True)
            crawler_task.add_log(f"结果公示: 新增 {list_count} 条")

            if not crawler_task._stop_flag:
                crawler_task.add_log(f"列表抓取完成，共新增 {list_count} 条，开始抓详情...")
                crawler_task.progress = {"phase": "detail", "detail": f"列表完成(新增{list_count}条), 开始抓详情...", "list_count": list_count}

                detail_count = await crawler.crawl_details(announcement_type=4)
                crawler_task.add_log(f"详情抓取完成: {detail_count} 条")

                # 删除无中标人的记录
                removed = crawler.remove_no_winner()
                crawler_task.add_log(f"清理无中标记录: 删除 {removed} 条")

                conn = sqlite3.connect(DB_PATH)
                total = conn.execute("SELECT COUNT(*) FROM announcements").fetchone()[0]
                with_winner = conn.execute("SELECT COUNT(*) FROM announcements WHERE winning_bidders IS NOT NULL AND winning_bidders != '[]'").fetchone()[0]
                conn.close()

                crawler_task.result = {
                    "status": "success",
                    "list_count": list_count,
                    "detail_count": detail_count,
                    "removed_no_winner": removed,
                    "total_records": total,
                    "with_winner": with_winner,
                }
                crawler_task.add_log(f"按时间采集完成! 新增={list_count}, 总记录={total}, 有中标={with_winner}")
            else:
                crawler_task.result = {"status": "stopped", "list_count": list_count}

        elif mode == "detail_only":
            # 仅抓详情（结果公示的）
            crawler_task.add_log("仅抓取结果公示详情页...")
            crawler_task.progress = {"phase": "detail", "detail": "抓取详情页..."}

            detail_count = await crawler.crawl_details(announcement_type=4)
            crawler_task.add_log(f"详情抓取完成: {detail_count} 条")

            # 删除无中标人的记录
            removed = crawler.remove_no_winner()
            crawler_task.add_log(f"清理无中标记录: 删除 {removed} 条")

            conn = sqlite3.connect(DB_PATH)
            total = conn.execute("SELECT COUNT(*) FROM announcements").fetchone()[0]
            with_winner = conn.execute("SELECT COUNT(*) FROM announcements WHERE winning_bidders IS NOT NULL AND winning_bidders != '[]'").fetchone()[0]
            conn.close()

            crawler_task.result = {
                "status": "success",
                "detail_count": detail_count,
                "removed_no_winner": removed,
                "total_records": total,
                "with_winner": with_winner,
            }
            crawler_task.add_log(f"详情采集完成! 总记录={total}, 有中标={with_winner}")

    finally:
        await crawler.close()


# ========== 响应模型 ==========

class WinningBidder(BaseModel):
    supplier_name: str | None = None
    bid_amount: float | None = None
    is_winning: int | None = None
    rank: int | None = None
    social_credit_code: str | None = None


class AnnouncementItem(BaseModel):
    id: str
    project_id: int | None = None
    project_no: str | None = None
    title: str
    announcement_type: int | None = None
    announcement_type_desc: str | None = None
    tender_mode_desc: str | None = None
    category: str | None = None
    publish_date: str | None = None
    deadline: str | None = None
    source_url: str | None = None
    # 深度字段
    purchase_control_price: float | None = None
    bid_price: float | None = None
    winning_bidders: list[WinningBidder] | None = None
    detail_fetched: int = 0
    tenderer_name: str | None = None
    tenderer_contact: str | None = None
    tenderer_phone: str | None = None
    project_address: str | None = None
    fund_source: str | None = None
    # 展开中标人后的字段
    current_bidder: WinningBidder | None = None  # 当前行对应的中标人
    bidder_index: int = 0  # 当前行是第几个中标人(0-based)
    bidder_total: int = 1  # 该公告共有几个中标人


class AnnouncementListResponse(BaseModel):
    total: int
    page: int
    size: int
    items: list[AnnouncementItem]


class StatItem(BaseModel):
    name: str
    value: int


class TrendItem(BaseModel):
    date: str
    count: int


class TopCompany(BaseModel):
    rank: int = 0
    name: str
    count: int = 0
    total_amount: float = 0.0

class DashboardStats(BaseModel):
    total: int
    today: int
    this_week: int
    this_month: int
    by_category: list[StatItem]
    by_type: list[StatItem]
    by_tender_mode: list[StatItem]
    daily_trend: list[TrendItem]
    total_bid_amount: float
    winning_count: int
    top_count_companies: list[TopCompany] = []  # 中标数最多公司排名
    top_amount_companies: list[TopCompany] = []  # 中标总金额最高公司排名


class CrawlerStartRequest(BaseModel):
    mode: str = "incremental"  # incremental / full / by_date / detail_only
    max_pages: int = 100
    days: int | None = None  # 按时间模式：最近N天


# ========== 接口 ==========

@app.get("/api/dashboard", response_model=DashboardStats)
async def get_dashboard_stats() -> DashboardStats:
    """获取Dashboard统计数据 - 按中标人(仅is_winning=1)展开计算"""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM announcements").fetchall()

    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    # 展开后的统计（只展开is_winning=1的中标人）
    today_count = 0
    week_count = 0
    month_count = 0
    category_map: dict[str, int] = {}
    type_map: dict[str, int] = {}
    mode_map: dict[str, int] = {}
    daily_map: dict[str, int] = {}
    total_bid = 0.0
    total_expanded = 0  # 展开后的总行数
    winning_count = 0

    # 公司统计: {公司名: {count, total_amount}}
    company_stats: dict[str, dict] = {}

    for r in rows:
        pd_str = r["publish_date"] or ""
        try:
            pd = datetime.fromisoformat(pd_str) if pd_str else None
            if pd and pd.tzinfo:
                pd = pd.replace(tzinfo=None)
        except (ValueError, TypeError):
            pd = None

        # 解析中标人，只取is_winning=1
        wbs_all = []
        wbs_str = r["winning_bidders"]
        if wbs_str and wbs_str != "[]":
            try:
                wbs_all = json.loads(wbs_str)
            except Exception:
                pass

        # 只保留中标(is_winning=1)的中标人
        wbs = [w for w in wbs_all if w.get("is_winning") == 1]

        row_count = len(wbs) if wbs else (1 if not wbs_all else 0)  # 只有中标人才展开
        total_expanded += row_count

        if pd:
            if pd >= today_start:
                today_count += row_count
            if pd >= week_ago:
                week_count += row_count
            if pd >= month_ago:
                month_count += row_count
            day_key = pd.strftime("%Y-%m-%d")
            daily_map[day_key] = daily_map.get(day_key, 0) + row_count

        cat = r["category"] or "未知"
        category_map[cat] = category_map.get(cat, 0) + row_count

        tdesc = r["announcement_type_desc"] or "未知"
        type_map[tdesc] = type_map.get(tdesc, 0) + row_count

        mode = r["tender_mode_desc"] or "未知"
        mode_map[mode] = mode_map.get(mode, 0) + row_count

        # 金额 + 公司统计（只统计中标人）
        for w in wbs:
            amt = w.get("bid_amount")
            if amt:
                total_bid += float(amt)
            winning_count += 1
            # 公司统计
            supplier = w.get("supplier_name", "")
            if supplier:
                if supplier not in company_stats:
                    company_stats[supplier] = {"count": 0, "total_amount": 0.0}
                company_stats[supplier]["count"] += 1
                if amt:
                    company_stats[supplier]["total_amount"] += float(amt)

    conn.close()

    by_category = [StatItem(name=k, value=v) for k, v in sorted(category_map.items(), key=lambda x: -x[1])]
    by_type = [StatItem(name=k, value=v) for k, v in sorted(type_map.items(), key=lambda x: -x[1])]
    by_tender_mode = [StatItem(name=k, value=v) for k, v in sorted(mode_map.items(), key=lambda x: -x[1])]
    daily_trend = [TrendItem(date=k, count=v) for k, v in sorted(daily_map.items(), reverse=True)[:30]]

    # 中标数最多公司排名 TOP10
    top_count_companies = []
    if company_stats:
        sorted_by_count = sorted(company_stats.items(), key=lambda x: -x[1]["count"])[:10]
        for rank, (name, stats) in enumerate(sorted_by_count, 1):
            top_count_companies.append(TopCompany(
                rank=rank, name=name, count=stats["count"], total_amount=stats["total_amount"],
            ))

    # 中标总金额最高公司排名 TOP10
    top_amount_companies = []
    if company_stats:
        sorted_by_amount = sorted(company_stats.items(), key=lambda x: -x[1]["total_amount"])[:10]
        for rank, (name, stats) in enumerate(sorted_by_amount, 1):
            top_amount_companies.append(TopCompany(
                rank=rank, name=name, count=stats["count"], total_amount=stats["total_amount"],
            ))

    return DashboardStats(
        total=total_expanded,
        today=today_count,
        this_week=week_count,
        this_month=month_count,
        by_category=by_category,
        by_type=by_type,
        by_tender_mode=by_tender_mode,
        daily_trend=daily_trend,
        total_bid_amount=total_bid,
        winning_count=winning_count,
        top_count_companies=top_count_companies,
        top_amount_companies=top_amount_companies,
    )


@app.get("/api/announcements", response_model=AnnouncementListResponse)
async def list_announcements(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    keyword: str | None = Query(None),
    category: str | None = Query(None),
    tender_mode: str | None = Query(None),
    announcement_type: int | None = Query(None),
    has_winner: bool | None = Query(None, description="是否有中标人"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
) -> AnnouncementListResponse:
    """查询公告列表 - 按中标人展开，每个中标人一行"""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM announcements ORDER BY publish_date DESC").fetchall()
    conn.close()

    filtered = list(rows)
    if keyword:
        kw = keyword.lower()
        def _match_keyword(r):
            # 搜索公告标题、项目编号、招标人、中标人名称
            if kw in (r["title"] or "").lower(): return True
            if kw in (r["project_no"] or "").lower(): return True
            if kw in (r["tenderer_name"] or "").lower(): return True
            # 搜索中标人名称
            if r["winning_bidders"]:
                try:
                    for wb in json.loads(r["winning_bidders"]):
                        if kw in (wb.get("supplier_name") or "").lower(): return True
                except Exception:
                    pass
            return False
        filtered = [r for r in filtered if _match_keyword(r)]
    if category:
        filtered = [r for r in filtered if r["category"] == category]
    if tender_mode:
        filtered = [r for r in filtered if r["tender_mode_desc"] == tender_mode]
    if announcement_type:
        filtered = [r for r in filtered if r["announcement_type"] == announcement_type]
    if has_winner:
        filtered = [r for r in filtered if r["winning_bidders"] and r["winning_bidders"] != "[]"]
    if start_date:
        filtered = [r for r in filtered if (r["publish_date"] or "") >= start_date]
    if end_date:
        filtered = [r for r in filtered if (r["publish_date"] or "") <= end_date + "T23:59:59"]

    # 按中标人展开：只展开中标(is_winning=1)的，每个中标人一行
    expanded = []
    for r in filtered:
        wbs_all = []
        if r["winning_bidders"]:
            try:
                wbs_all = [WinningBidder(**w) for w in json.loads(r["winning_bidders"])]
            except Exception:
                pass

        # 只保留中标(is_winning=1)的中标人
        wbs = [wb for wb in wbs_all if wb.is_winning == 1]

        if not wbs:
            # 没有中标人的，保持一行
            expanded.append(AnnouncementItem(
                id=r["id"],
                project_id=r["project_id"],
                project_no=r["project_no"],
                title=r["title"] or "",
                announcement_type=r["announcement_type"],
                announcement_type_desc=r["announcement_type_desc"],
                tender_mode_desc=r["tender_mode_desc"],
                category=r["category"],
                publish_date=r["publish_date"],
                deadline=r["deadline"],
                source_url=r["source_url"],
                purchase_control_price=r["purchase_control_price"],
                bid_price=r["bid_price"],
                winning_bidders=[],
                detail_fetched=r["detail_fetched"],
                tenderer_name=r["tenderer_name"],
                tenderer_contact=r["tenderer_contact"],
                tenderer_phone=r["tenderer_phone"],
                project_address=r["project_address"],
                fund_source=r["fund_source"],
                current_bidder=None,
                bidder_index=0,
                bidder_total=1,
            ))
        else:
            # 有中标人的，展开为n行（只展开中标的）
            for idx, wb in enumerate(wbs):
                expanded.append(AnnouncementItem(
                    id=r["id"],
                    project_id=r["project_id"],
                    project_no=r["project_no"],
                    title=r["title"] or "",
                    announcement_type=r["announcement_type"],
                    announcement_type_desc=r["announcement_type_desc"],
                    tender_mode_desc=r["tender_mode_desc"],
                    category=r["category"],
                    publish_date=r["publish_date"],
                    deadline=r["deadline"],
                    source_url=r["source_url"],
                    purchase_control_price=r["purchase_control_price"],
                    bid_price=r["bid_price"],
                    winning_bidders=wbs,
                    detail_fetched=r["detail_fetched"],
                    tenderer_name=r["tenderer_name"],
                    tenderer_contact=r["tenderer_contact"],
                    tenderer_phone=r["tenderer_phone"],
                    project_address=r["project_address"],
                    fund_source=r["fund_source"],
                    current_bidder=wb,
                    bidder_index=idx,
                    bidder_total=len(wbs),
                ))

    total = len(expanded)
    start_idx = (page - 1) * size
    page_items = expanded[start_idx:start_idx + size]

    return AnnouncementListResponse(total=total, page=page, size=size, items=page_items)


@app.get("/api/announcements/{announcement_id}")
async def get_announcement(announcement_id: str) -> dict[str, Any]:
    """获取公告详情"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM announcements WHERE id = ?", (announcement_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="公告不存在")
    result = dict(row)
    for field in ["winning_bidders", "candidate_info", "change_records", "raw_list_data", "raw_detail_data"]:
        if result.get(field):
            try:
                result[field] = json.loads(result[field])
            except Exception:
                pass
    return result


@app.post("/api/crawler/start")
async def start_crawler(req: CrawlerStartRequest = CrawlerStartRequest()):
    """启动采集数据任务（真正执行）"""
    if crawler_task.is_running:
        return {
            "status": "already_running",
            "message": f"采集数据正在运行中 (mode={crawler_task.mode})",
            "task": crawler_task.to_dict(),
        }

    task_id = f"task_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    crawler_task.task_id = task_id
    crawler_task.logs = []

    # 在后台线程中运行采集数据
    thread = threading.Thread(
        target=_run_crawler_sync,
        args=(req.mode, req.max_pages, req.days),
        daemon=True,
    )
    crawler_task._thread = thread
    thread.start()

    mode_label = {"incremental": "增量", "full": "全量", "by_date": "按时间", "detail_only": "仅详情"}.get(req.mode, req.mode)
    extra = f", 最近{req.days}天" if req.mode == "by_date" and req.days else ""
    return {
        "status": "started",
        "task_id": task_id,
        "message": f"{mode_label}采集已启动 (max_pages={req.max_pages}{extra})",
    }


@app.post("/api/crawler/stop")
async def stop_crawler():
    """停止采集数据任务"""
    if not crawler_task.is_running:
        return {"status": "not_running", "message": "采集数据未在运行"}
    crawler_task._stop_flag = True
    crawler_task.add_log("收到停止信号...", "warning")
    return {"status": "stopping", "message": "正在停止采集数据..."}


@app.get("/api/crawler/status")
async def get_crawler_status():
    """获取采集数据任务状态"""
    return crawler_task.to_dict()


@app.get("/api/categories")
async def get_categories() -> list[str]:
    conn = get_conn()
    rows = conn.execute("SELECT DISTINCT category FROM announcements WHERE category IS NOT NULL").fetchall()
    conn.close()
    return sorted([r[0] for r in rows])


@app.get("/api/tender-modes")
async def get_tender_modes() -> list[str]:
    conn = get_conn()
    rows = conn.execute("SELECT DISTINCT tender_mode_desc FROM announcements WHERE tender_mode_desc IS NOT NULL").fetchall()
    conn.close()
    return sorted([r[0] for r in rows])


@app.get("/api/announcement-types")
async def get_announcement_types() -> list[dict]:
    return [
        {"value": 1, "label": "招标公告"},
        {"value": 2, "label": "变更公告"},
        {"value": 3, "label": "候选人公示"},
        {"value": 4, "label": "结果公示"},
        {"value": 5, "label": "邀请函"},
    ]


@app.get("/api/health")
async def health_check() -> dict[str, Any]:
    conn = get_conn()
    total = conn.execute("SELECT COUNT(*) FROM announcements").fetchone()[0]
    with_detail = conn.execute("SELECT COUNT(*) FROM announcements WHERE detail_fetched = 1").fetchone()[0]
    with_winner = conn.execute("SELECT COUNT(*) FROM announcements WHERE winning_bidders IS NOT NULL AND winning_bidders != '[]'").fetchone()[0]
    conn.close()
    return {
        "status": "ok",
        "total_records": total,
        "with_detail": with_detail,
        "with_winner": with_winner,
        "crawler_running": crawler_task.is_running,
    }


# ========== 导出接口 ==========

def _query_export_data(
    keyword: str | None = None,
    category: str | None = None,
    tender_mode: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[dict]:
    """查询导出数据（固定：结果公示+有中标）"""
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM announcements WHERE announcement_type = 4 AND winning_bidders IS NOT NULL AND winning_bidders != '[]' ORDER BY publish_date DESC"
    ).fetchall()
    conn.close()

    filtered = list(rows)
    if keyword:
        kw = keyword.lower()
        def _match_keyword(r):
            if kw in (r["title"] or "").lower(): return True
            if kw in (r["project_no"] or "").lower(): return True
            if kw in (r["tenderer_name"] or "").lower(): return True
            if r["winning_bidders"]:
                try:
                    for wb in json.loads(r["winning_bidders"]):
                        if kw in (wb.get("supplier_name") or "").lower(): return True
                except Exception:
                    pass
            return False
        filtered = [r for r in filtered if _match_keyword(r)]
    if category:
        filtered = [r for r in filtered if r["category"] == category]
    if tender_mode:
        filtered = [r for r in filtered if r["tender_mode_desc"] == tender_mode]
    if start_date:
        filtered = [r for r in filtered if (r["publish_date"] or "") >= start_date]
    if end_date:
        filtered = [r for r in filtered if (r["publish_date"] or "") <= end_date + "T23:59:59"]

    result = []
    for r in filtered:
        wbs_all = []
        if r["winning_bidders"]:
            try:
                wbs_all = json.loads(r["winning_bidders"])
            except Exception:
                pass

        # 只保留中标(is_winning=1)的中标人
        wbs = [w for w in wbs_all if w.get("is_winning") == 1]

        if not wbs:
            # 没有中标人的，保持一行
            result.append({
                "id": r["id"],
                "project_no": r["project_no"] or "",
                "title": r["title"] or "",
                "category": r["category"] or "",
                "tender_mode_desc": r["tender_mode_desc"] or "",
                "tenderer_name": r["tenderer_name"] or "",
                "tenderer_contact": r["tenderer_contact"] or "",
                "tenderer_phone": r["tenderer_phone"] or "",
                "publish_date": r["publish_date"] or "",
                "project_address": r["project_address"] or "",
                "fund_source": r["fund_source"] or "",
                "purchase_control_price": r["purchase_control_price"] or "",
                "bid_price": r["bid_price"] or "",
                "winner_supplier": "",
                "winner_amount": "",
                "winner_credit_code": "",
                "source_url": r["source_url"] or "",
            })
        else:
            # 每个中标人一行（只展开中标的）
            for w in wbs:
                result.append({
                    "id": r["id"],
                    "project_no": r["project_no"] or "",
                    "title": r["title"] or "",
                    "category": r["category"] or "",
                    "tender_mode_desc": r["tender_mode_desc"] or "",
                    "tenderer_name": r["tenderer_name"] or "",
                    "tenderer_contact": r["tenderer_contact"] or "",
                    "tenderer_phone": r["tenderer_phone"] or "",
                    "publish_date": r["publish_date"] or "",
                    "project_address": r["project_address"] or "",
                    "fund_source": r["fund_source"] or "",
                    "purchase_control_price": r["purchase_control_price"] or "",
                    "bid_price": r["bid_price"] or "",
                    "winner_supplier": w.get("supplier_name", ""),
                    "winner_amount": w.get("bid_amount", ""),
                    "winner_credit_code": w.get("social_credit_code", ""),
                    "source_url": r["source_url"] or "",
                })
    return result


@app.get("/api/export/csv")
async def export_csv(
    keyword: str | None = Query(None),
    category: str | None = Query(None),
    tender_mode: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
):
    """导出中标结果为CSV"""
    data = _query_export_data(keyword, category, tender_mode, start_date, end_date)

    output = io.StringIO()
    # UTF-8 BOM for Excel
    output.write("\ufeff")

    headers = [
        "公告ID", "项目编号", "公告标题", "分类", "招标方式", "招标人",
        "联系人", "联系电话",
        "发布时间", "项目地址", "资金来源",
        "控制价", "中标金额", "中标人", "中标人信用代码", "详情链接",
    ]
    writer = csv.writer(output)
    writer.writerow(headers)
    for d in data:
        writer.writerow([
            d["id"], d["project_no"], d["title"], d["category"],
            d["tender_mode_desc"], d["tenderer_name"],
            d["tenderer_contact"], d["tenderer_phone"],
            d["publish_date"],
            d["project_address"], d["fund_source"],
            d["purchase_control_price"], d["winner_amount"],
            d["winner_supplier"],
            d["winner_credit_code"], d["source_url"],
        ])

    filename = f"中标结果_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@app.get("/api/export/excel")
async def export_excel(
    keyword: str | None = Query(None),
    category: str | None = Query(None),
    tender_mode: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
):
    """导出中标结果为Excel"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    data = _query_export_data(keyword, category, tender_mode, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "中标结果"

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    amount_font = Font(name="微软雅黑", size=11, color="CC0000", bold=True)
    winner_font = Font(name="微软雅黑", size=11, color="1677FF")
    thin_border = Border(
        left=Side(style="thin", color="E8E8E8"),
        right=Side(style="thin", color="E8E8E8"),
        top=Side(style="thin", color="E8E8E8"),
        bottom=Side(style="thin", color="E8E8E8"),
    )

    headers = [
        ("公告ID", 10), ("项目编号", 18), ("公告标题", 45), ("分类", 8),
        ("招标方式", 12), ("招标人", 20), ("联系人", 10), ("联系电话", 14),
        ("发布时间", 18),
        ("项目地址", 25), ("资金来源", 15), ("控制价", 14), ("中标金额", 14),
        ("中标人", 22), ("中标人信用代码", 22), ("详情链接", 30),
    ]

    # 写表头
    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 写数据
    for row_idx, d in enumerate(data, 2):
        values = [
            d["id"], d["project_no"], d["title"], d["category"],
            d["tender_mode_desc"], d["tenderer_name"],
            d["tenderer_contact"], d["tenderer_phone"],
            d["publish_date"],
            d["project_address"], d["fund_source"],
            d["purchase_control_price"], d["winner_amount"],
            d["winner_supplier"],
            d["winner_credit_code"], d["source_url"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
            cell.alignment = cell_align
            cell.border = thin_border
            # 金额列高亮 (12=控制价, 13=中标金额)
            if col_idx in (12, 13) and val:
                cell.font = amount_font
                cell.number_format = "#,##0.00"
            # 中标人列高亮 (14=中标人)
            if col_idx == 14 and val:
                cell.font = winner_font

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(data) + 1}"

    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"中标结果_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
